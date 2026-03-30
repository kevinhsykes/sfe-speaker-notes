import streamlit as st
import json
import os
import io
from datetime import datetime

# --- Config ---
st.set_page_config(page_title="SFE Speaker Notes", page_icon="🎤", layout="wide")
DATA_FILE = "sfe_data.json"

# --- Universal Questions (asked to every speaker for cross-comparison) ---
UNIVERSAL_QUESTIONS = [
    {"id": "readiness", "question": "On a 1-10 scale, how ready is your organization to adopt AI?", "type": "quantitative"},
    {"id": "workflow_pct", "question": "What percentage of your daily workflows currently involve AI?", "type": "quantitative"},
    {"id": "tool_count", "question": "How many AI tools does your organization actively use?", "type": "quantitative"},
    {"id": "weekly_users", "question": "How many employees at your org use AI weekly?", "type": "quantitative"},
    {"id": "ai_journey", "question": "Where are you in the AI journey: not started / exploring / piloting / scaling / mature?", "type": "categorical"},
    {"id": "consultant_fix", "question": "If you hired an AI consultant tomorrow, what would you ask them to fix first?", "type": "categorical"},
]

# --- Default Speaker Data ---
DEFAULT_SPEAKERS = [
    {
        "id": "connelly",
        "name": "Lindsay Connelly",
        "date": "2026-03-18",
        "company": "Linden Digital Marketing",
        "role": "Founder",
        "task_ref": "Task 1b (running doc, first entry)",
        "bio": "Founder of Linden Digital Marketing, a Rochester-area digital marketing agency.",
        "company_info": {
            "industry": "Digital Marketing",
            "location": "Rochester, NY",
            "size": "Small business / agency",
            "founded": "",
            "website": "",
            "notes": "Full-service digital marketing agency serving Rochester-area businesses. Services likely include SEO, social media management, content creation, paid advertising, and analytics."
        },
        "has_conversation_task": False,
        "ai_usage": [
            {"prompt": "AI tools used in her digital marketing work (content generation, analytics, ad targeting, etc.)", "notes": "Uses AI for email outreach/marketing automation. Has an AI-driven pricing algorithm that generates ROI predictions for client campaigns. Also uses a lead qualification program to help decide which leads are worth pursuing."},
            {"prompt": "How she positions AI capabilities to clients", "notes": "Very open about AI with clients, frames it as a major factor in digital marketing. Uses a numbers-first approach, has a table of ROI data she references when closing deals. Shows clients exactly what return they'd get for the cost of her campaigns."},
            {"prompt": "What clients ask for vs. what actually works", "notes": "Tailors each client to the correct advertising platform based on years of hands-on social media experience. Backs everything with measurable ROI data. The numbers close the deals, she knows which platforms work for which clients and can demonstrate concrete value."},
            {"prompt": "Any tools or platforms she relies on day to day", "notes": "Email AI tools for organizing/summarizing, lead qualification program, ROI/pricing algorithm. Specific tool names not captured."},
            {"prompt": "Her take on AI replacing vs. augmenting marketing roles", "notes": "Very bullish, early adopter. Jumped on AI immediately. Started by using it for her own email workflow (organizing, summarizing) before expanding into client-facing applications like the pricing algorithm and lead qualification."}
        ],
        "takeaways": [
            {"prompt": "How she started Linden Digital Marketing", "notes": "Started the business on her own. At RIT she founded a rock climbing club, which was a major life-changing moment for her. That's where she learned she enjoyed leading others and that she was good at it. That leadership instinct carried into starting Linden."},
            {"prompt": "Biggest challenges running a small marketing agency", "notes": ""},
            {"prompt": "Advice for someone entering consulting/agency work", "notes": ""},
            {"prompt": "Anything about Rochester's small business market", "notes": ""},
            {"prompt": "Most interesting or unexpected thing she said", "notes": "Her numbers-first approach. The ROI table she uses when closing deals, the pricing algorithm for predicting client returns, the emphasis on measurable outcomes. She knew her numbers, knew the social media platforms from years of practice, and could show clients exactly what they'd get. The combination of deep platform knowledge and data-backed selling was the strongest part of the presentation."},
            {"prompt": "Team culture / hiring philosophy", "notes": "Her first hire was a college roommate. She hired another close friend later on. Friends working together isn't always something to avoid, if you know you work well together it can be a real strength. The team environment, support, and camaraderie was a major takeaway from the talk. When one of her employees got cancer, the whole team rallied around her. Lindsay gave her a raise to help her get through it. Great for morale all around, speaks to the kind of environment she's built."}
        ],
        "comparisons": [
            {"prompt": "How does a marketing agency's AI adoption compare to other speakers?", "notes": ""},
            {"prompt": "Service business vs. product business differences", "notes": ""}
        ],
        "prepared_questions": [],
        "conversation_notes": "No post-talk conversation. She was swarmed after the talk.",
        "conversation_takeaways": "",
        "raw_shorthand": "Wasn't prepared for this speaker, minimal notes taken. Backfilled from memory."
    },
    {
        "id": "weis",
        "name": "Cyndi Weis",
        "date": "2026-03-26",
        "company": "Breathe Yoga & Juice Bar",
        "role": "Founder & Owner",
        "task_ref": "Task 1b (running doc)",
        "bio": "20+ year wellness business in Pittsford. Yoga studio, juice bar, kitchen/bakery, spa, retail boutique. Family-run with daughters Abby and Carly. 2016 Rochester Business Person of the Year (50+ employees). Previously had 5 franchise locations, dissolved franchise model in 2022. Registered dietitian and keynote speaker on aging/longevity.",
        "company_info": {
            "industry": "Wellness / Hospitality",
            "location": "Pittsford, NY (Rochester suburb)",
            "size": "50+ employees at peak, family-run",
            "founded": "~2004",
            "website": "breatheyogabar.com",
            "notes": "Multi-revenue-stream wellness business: yoga studio, juice bar, kitchen/bakery, spa, retail boutique. Previously franchised to 5 locations, dissolved franchise model in 2022. Cyndi is a registered dietitian and keynote speaker on aging/longevity. Adapted during COVID with virtual classes, curbside service, and a 40 Days wellness program."
        },
        "has_conversation_task": False,
        "ai_usage": [
            {"prompt": "Does she use any AI tools currently? (scheduling, social media, inventory, content)", "notes": "Uses one consolidated system across all four business units for payroll and operations (name of system not recalled). No specific AI tools mentioned during the talk or conversation."},
            {"prompt": "What operational tasks eat the most time that tech could help with?", "notes": "Understanding customer churn. She wants to know why people cancel, why they stop showing up. Email follow-ups work for class/spa cancellations where she has membership data, but the retail and juice bar side is much harder to track since those customers can just stop coming without any formal cancellation."},
            {"prompt": "Attitude toward AI: open/skeptical/unaware?", "notes": "Curious and quick to adapt. COVID proved she can pivot to technology fast. Took her entire yoga class offering online immediately and it stuck, still does a significant percentage of classes online today (possibly around 30%, exact number not recalled). Not resistant to tech at all."},
            {"prompt": "How does she use customer data across revenue streams (classes, food, spa, retail)?", "notes": "Has better visibility into class and spa customers (membership/booking data) than retail and juice bar customers (transactional, no membership). Her biggest question is understanding the group she loses, why people cancel or stop coming across all the different parts of the business."},
            {"prompt": "If an AI consultant pitched her, what would she actually want help with?", "notes": "Understanding customer churn. Why the cancellations happen, what the group she doesn't retain looks like. Why someone stopped buying smoothies, why they stopped coming to yoga, why they didn't buy the jacket. She doesn't have great information on this, especially on the retail/juice bar side."},
            {"prompt": "Any mention of online class platform, virtual offerings, and tech behind that", "notes": "Went online immediately when COVID hit. All members got instant access to online yoga classes. Made a deliberate choice that every instructor films in front of a plain white wall, nothing to distract from pure yoga instruction. Online classes continued post-COVID and still make up a significant share of her yoga business (estimated around 30%, exact figure not recalled)."}
        ],
        "takeaways": [
            {"prompt": "Lessons from scaling to 5 franchises then dissolving the franchise model", "notes": "Big part of the talk. She chose to disenfranchise because she wanted full control and for everything to fit her vision. Offered franchisees the option to retain their businesses, but they had to change their names and operate independently. The logistical headache was gift cards and coupons. Around July of the transition year, her team had to reach out to everyone who had purchased gift cards at franchise locations to communicate that they would expire at the end of the year and could only be used at the location where they were purchased. After the new year when disenfranchisement was finalized, they would no longer be valid at any Breathe location. Required repeated communication to make sure customers didn't feel wronged. Very stressful process."},
            {"prompt": "How she manages multiple revenue streams under one roof", "notes": "Uses one consolidated system for payroll and operations across all four units (yoga, juice bar, spa, retail). System name not recalled."},
            {"prompt": "Customer retention in wellness (people cycle in and out)", "notes": "Retention is a major strength. Her introductory offer is intentionally different from competitors, no '2 free classes' model. Instead, new customers pay around $30 for two weeks of unlimited, all-access yoga. Her reasoning is that one or two classes is not enough for meaningful life changes, so she gives the full experience during the trial. Once customers are in, she rarely loses them. Yoga becomes a lifestyle change that lasts. Even customers who have to step away tend to come back eventually because the value sticks. Retention rate is high (specific numbers not recalled). Her biggest open question is understanding the customers she does lose, why they cancel, why they stop coming."},
            {"prompt": "Running a family business with her daughters", "notes": ""},
            {"prompt": "Building around personal passion vs. market demand", "notes": "Work isn't work if you love it. Her husband told her it was going to be a ton of work, long hours, tiring. She already knew that, did it anyway, and has no regrets."},
            {"prompt": "How she adapted during COVID (virtual classes, curbside, 40 Days program)", "notes": "Moved yoga classes online immediately. All members got instant access. Instructors always filmed in front of a plain white wall for clean, distraction-free instruction. Online classes stuck and still represent a meaningful share of the business. The side door entrance is actually a longstanding design choice from day one, not a COVID pivot. She wanted the place to feel hole-in-the-wall, exclusive, homely, rather than a big storefront with signage on main street."}
        ],
        "comparisons": [
            {"prompt": "Small wellness business vs. Connelly's digital marketing agency: different AI needs?", "notes": "Connelly is already using AI actively in her work (email automation, ROI prediction, lead qualification). Weis is pre-AI, no specific tools mentioned. Different stages of the journey entirely."},
            {"prompt": "Owner-operator perspective vs. more corporate/scaled businesses (Mucci, Goldner)", "notes": ""},
            {"prompt": "How does a brick-and-mortar service business think about tech differently?", "notes": "Quick to adopt when forced (COVID), but the ongoing challenge is that brick-and-mortar service businesses generate messier, more fragmented customer data than digital-first businesses. Connelly's marketing agency has clean digital data by nature of the work. Weis has four physical revenue streams with different customer tracking depending on whether it's membership-based (yoga, spa) or transactional (juice bar, retail)."}
        ],
        "prepared_questions": [],
        "conversation_notes": "Spoke with Cyndi for about 10 minutes after the talk. The number one thing she wanted to know is why the people who cancel her services leave. Why the cancellations, what is the group she doesn't retain. Why didn't they buy that jacket, why did they stop coming in to buy smoothies, why did they stop coming to yoga class. She doesn't have great information on it. Email follow-ups work when people stop classes or spa treatments, but it is harder to get that information on the retail/juice side where people just stop making purchases without any formal cancellation.",
        "conversation_takeaways": "Her core problem is a data/visibility gap around customer churn. She needs better ways to capture why people leave across all four business units, especially the transactional ones where there's no membership to cancel.",
        "raw_shorthand": "Backfilled from memory. Did not have a structured shorthand system ready for this talk.",
        "universal_answers": {"consultant_fix": "Wants to understand why customers cancel/leave. Specifically, why people stop coming to yoga, stop buying smoothies, stop purchasing retail. Has some ability to follow up via email when members cancel classes or spa treatments, but lacks good data on the retail/juice bar side where drop-off is harder to track."}
    },
    {
        "id": "goldner",
        "name": "Andrew Goldner",
        "date": "2026-04-01",
        "company": "(Research before talk)",
        "role": "(Research before talk)",
        "task_ref": "Task 1b (running doc) + Learning Contract Task 3 (pursue conversation)",
        "bio": "Learning Contract Task 3 focus: consulting, business growth strategy, where AI fits into the advisory model. Also use this conversation to research possible data sources/ideas for Learning Contract Task 8 (applied analytics project).",
        "company_info": {
            "industry": "Consulting / Advisory (confirm)",
            "location": "(Research)",
            "size": "(Research)",
            "founded": "(Research)",
            "website": "(Research)",
            "notes": "Need to research before April 1 talk. Focus areas: VC/consulting background, business growth strategy, advisory model."
        },
        "has_conversation_task": True,
        "ai_usage": [
            {"prompt": "How AI fits into his consulting/advisory model", "notes": ""},
            {"prompt": "What AI tools his firm uses internally", "notes": ""},
            {"prompt": "How he advises clients on AI adoption", "notes": ""},
            {"prompt": "Where he sees AI creating the most value in consulting", "notes": ""},
            {"prompt": "His take on AI implementation failure rates", "notes": ""}
        ],
        "takeaways": [
            {"prompt": "His path into consulting/advisory work", "notes": ""},
            {"prompt": "How he thinks about business growth strategy", "notes": ""},
            {"prompt": "What skills matter most in consulting", "notes": ""},
            {"prompt": "Advice for someone entering the field", "notes": ""}
        ],
        "comparisons": [
            {"prompt": "Advisory/consulting perspective vs. business operators (Weis, Connelly)", "notes": ""},
            {"prompt": "How does his view on AI adoption differ from someone implementing it internally?", "notes": ""}
        ],
        "prepared_questions": [
            "What does your consulting/advisory process look like end to end?",
            "Where does AI fit into how you advise clients?",
            "What data sources or datasets do you rely on for client analysis?",
            "What skills matter most for someone entering consulting?",
            "Any leads on datasets or project ideas for an applied analytics project?"
        ],
        "conversation_notes": "",
        "conversation_takeaways": "",
        "raw_shorthand": ""
    },
    {
        "id": "mucci",
        "name": "Martin Mucci",
        "date": "2026-04-08",
        "company": "Paychex",
        "role": "Former CEO",
        "task_ref": "Task 1b (running doc) + Learning Contract Task 5 (pursue conversation)",
        "bio": "Led Paychex as CEO during major growth period. Paychex founded by Tom Golisano in 1971 with $3,000 and a credit card. Major Rochester employer currently in the middle of a significant AI transformation.",
        "company_info": {
            "industry": "HR / Payroll / Financial Services",
            "location": "Rochester, NY (HQ)",
            "size": "~16,000 employees, Fortune 500",
            "founded": "1971 by Tom Golisano",
            "website": "paychex.com",
            "notes": "One of the largest payroll and HR services companies in the US. Currently undergoing major AI transformation. Tom Golisano (founder) also founded the Golisano Institute. Key Rochester employer. Learning Contract Task 5 focus: AI transformation at Paychex, adoption challenges, what skills companies look for in implementation roles."
        },
        "has_conversation_task": True,
        "ai_usage": [
            {"prompt": "Where Paychex is in its AI transformation journey", "notes": ""},
            {"prompt": "Biggest adoption challenges internally", "notes": ""},
            {"prompt": "How leadership communicates AI changes to the workforce", "notes": ""},
            {"prompt": "Specific AI tools or capabilities Paychex has rolled out", "notes": ""},
            {"prompt": "What went well vs. what stalled", "notes": ""},
            {"prompt": "How they measure success of AI initiatives", "notes": ""}
        ],
        "takeaways": [
            {"prompt": "His leadership philosophy at scale", "notes": ""},
            {"prompt": "Paychex's competitive position and where AI fits in that", "notes": ""},
            {"prompt": "Rochester's role in Paychex's future", "notes": ""},
            {"prompt": "Advice for someone early in their career", "notes": ""}
        ],
        "comparisons": [
            {"prompt": "Enterprise-scale AI transformation vs. small business adoption (Weis, Connelly)", "notes": ""},
            {"prompt": "How does Paychex's approach compare to Lawley Insurance (Adam Clouden conversation)?", "notes": ""},
            {"prompt": "Leadership-driven vs. bottom-up adoption: what does Mucci's perspective add?", "notes": ""}
        ],
        "prepared_questions": [
            "What has been the biggest AI adoption challenge at Paychex?",
            "What does the internal change management process look like?",
            "What skills does Paychex look for in people helping with AI implementation?",
            "How involved is leadership vs. middle management in driving adoption?",
            "What would you want from an outside AI implementation consultant?",
            "How does Paychex handle employee resistance or skepticism?"
        ],
        "conversation_notes": "",
        "conversation_takeaways": "",
        "raw_shorthand": ""
    }
]

DEFAULT_ANALYSIS = {
    "common_themes": "",
    "biggest_differences": "",
    "consulting_insights": "From Cyndi Weis conversation: the most pressing problem for small businesses is connecting siloed data across business units before any AI layer can be applied. Tool selection is secondary to data infrastructure.",
    "surprises": ""
}

# --- Data Persistence ---
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {"speakers": DEFAULT_SPEAKERS, "analysis": DEFAULT_ANALYSIS}

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

# --- Initialize ---
data = load_data()
speakers = data["speakers"]
analysis = data.get("analysis", DEFAULT_ANALYSIS)

# --- Custom CSS ---
st.markdown("""
<style>
    .speaker-header {
        background: linear-gradient(135deg, #1B4F72, #2E86C1);
        padding: 20px 25px;
        border-radius: 10px;
        color: white;
        margin-bottom: 20px;
    }
    .speaker-header h2 { color: white !important; margin-bottom: 5px; }
    .speaker-header p { color: #D6EAF8; margin: 2px 0; }
    .company-card {
        background: #F8F9FA;
        border: 1px solid #DEE2E6;
        border-radius: 10px;
        padding: 16px 20px;
        margin-bottom: 20px;
    }
    .company-card h4 { margin-top: 0; color: #1B4F72; }
    .info-row { margin: 4px 0; font-size: 0.9em; }
    .info-label { font-weight: bold; color: #5D6D7E; }
    .prompt-box {
        background: #EBF5FB;
        border-left: 3px solid #2E86C1;
        padding: 10px 14px;
        border-radius: 0 6px 6px 0;
        margin: 8px 0 4px 0;
        font-size: 0.88em;
        color: #1B4F72;
    }
    .section-divider { border-top: 2px solid #E5E8E8; margin: 30px 0; }
    .status-upcoming {
        background: #FEF9E7; border: 1px solid #F9E79F;
        padding: 6px 14px; border-radius: 20px; font-size: 0.8em; color: #7D6608;
    }
    .status-completed {
        background: #E8F8F5; border: 1px solid #A3E4D7;
        padding: 6px 14px; border-radius: 20px; font-size: 0.8em; color: #1E8449;
    }
    .status-needs-notes {
        background: #FDEDEC; border: 1px solid #F5B7B1;
        padding: 6px 14px; border-radius: 20px; font-size: 0.8em; color: #922B21;
    }
</style>
""", unsafe_allow_html=True)

# --- Sidebar ---
st.sidebar.title("🎤 SFE Speaker Notes")
st.sidebar.markdown("*Spring 2026 | Kevin Sykes*")
st.sidebar.markdown("*Golisano Institute for Business & Entrepreneurship*")
st.sidebar.markdown("---")

nav_options = ["📊 Dashboard"]
for s in speakers:
    speaker_date = datetime.strptime(s["date"], "%Y-%m-%d")
    status = "✅" if speaker_date <= datetime.now() else "📅"
    nav_options.append(f"{status} {s['name']}")
nav_options.append("🔍 Cross-Speaker Analysis")
nav_options.append("➕ Add Speaker")

selection = st.sidebar.radio("Navigate", nav_options, label_visibility="collapsed")

st.sidebar.markdown("---")
st.sidebar.markdown("**Learning Contract References**")
st.sidebar.markdown("Task 1b: Running SFE notes")
st.sidebar.markdown("Task 3: Goldner conversation")
st.sidebar.markdown("Task 5: Mucci conversation")
st.sidebar.markdown("---")
st.sidebar.markdown("**Data Status**")
st.sidebar.markdown("⚠️ Data saves to server JSON.")
st.sidebar.markdown("May reset on app restart.")
st.sidebar.caption("Future: Google Sheets backend for permanent storage")

# --- Helper: count notes filled ---
def count_notes(speaker):
    filled = 0
    total = 0
    for section in ["ai_usage", "takeaways", "comparisons"]:
        for item in speaker.get(section, []):
            total += 1
            if item.get("notes", "").strip():
                filled += 1
    return filled, total

# --- Dashboard ---
if selection == "📊 Dashboard":
    st.title("SFE Speaker Notes Dashboard")
    st.markdown("Speaking from Experience, Spring 2026 | AI & Business Certificate")
    st.markdown("---")

    col1, col2, col3, col4 = st.columns(4)
    total = len(speakers)
    past = sum(1 for s in speakers if datetime.strptime(s["date"], "%Y-%m-%d") <= datetime.now())
    filled_speakers = 0
    total_notes_filled = 0
    total_notes_possible = 0
    for s in speakers:
        f, t = count_notes(s)
        total_notes_filled += f
        total_notes_possible += t
        if f > 0:
            filled_speakers += 1

    col1.metric("Total Speakers", total)
    col2.metric("Talks Attended", past)
    col3.metric("With Notes", filled_speakers)
    col4.metric("Notes Filled", f"{total_notes_filled}/{total_notes_possible}")

    st.markdown("---")
    st.subheader("Speaker Timeline")

    for s in speakers:
        speaker_date = datetime.strptime(s["date"], "%Y-%m-%d")
        is_past = speaker_date <= datetime.now()
        filled, total_prompts = count_notes(s)

        if is_past and filled > 0:
            pct = int((filled / total_prompts) * 100) if total_prompts > 0 else 0
            status_html = f'<span class="status-completed">✅ {filled}/{total_prompts} notes ({pct}%)</span>'
        elif is_past:
            status_html = '<span class="status-needs-notes">⚠️ Attended, needs notes</span>'
        else:
            days_until = (speaker_date - datetime.now()).days + 1
            status_html = f'<span class="status-upcoming">📅 In {days_until} days</span>'

        conv_tag = " 🗣️" if s.get("has_conversation_task") else ""

        st.markdown(f"""
        **{s['name']}** — {s['company']}{conv_tag}
        {speaker_date.strftime('%B %d, %Y')} &nbsp;&nbsp; {status_html}
        *{s['task_ref']}*
        """, unsafe_allow_html=True)
        st.markdown("---")

    # --- Export to Excel ---
    st.markdown("---")
    st.subheader("📥 Export to Excel")
    st.markdown("*Download all speaker data as an Excel file. Use as backup, for analysis, or to share with Jamarr.*")

    if st.button("Export All Data to Excel", type="primary"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()

            # --- Sheet 1: Speaker Overview ---
            ws1 = wb.active
            ws1.title = "Speaker Overview"
            headers = ["Speaker", "Company", "Role", "Date", "Task Ref", "Notes Filled", "Shorthand Notes"]
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")

            for col, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row, s in enumerate(speakers, 2):
                f, t = count_notes(s)
                ws1.cell(row=row, column=1, value=s["name"])
                ws1.cell(row=row, column=2, value=s["company"])
                ws1.cell(row=row, column=3, value=s["role"])
                ws1.cell(row=row, column=4, value=s["date"])
                ws1.cell(row=row, column=5, value=s["task_ref"])
                ws1.cell(row=row, column=6, value=f"{f}/{t}")
                ws1.cell(row=row, column=7, value=s.get("raw_shorthand", ""))

            for col in range(1, len(headers) + 1):
                ws1.column_dimensions[chr(64 + col)].width = 20

            # --- Sheet 2: Detailed Notes ---
            ws2 = wb.create_sheet("Detailed Notes")
            detail_headers = ["Speaker", "Section", "Prompt", "Notes"]
            for col, h in enumerate(detail_headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            detail_row = 2
            for s in speakers:
                for section_key, section_name in [("ai_usage", "AI Usage"), ("takeaways", "Key Takeaways"), ("comparisons", "Comparisons")]:
                    for item in s.get(section_key, []):
                        ws2.cell(row=detail_row, column=1, value=s["name"])
                        ws2.cell(row=detail_row, column=2, value=section_name)
                        ws2.cell(row=detail_row, column=3, value=item.get("prompt", ""))
                        ws2.cell(row=detail_row, column=4, value=item.get("notes", ""))
                        detail_row += 1

                if s.get("conversation_notes"):
                    ws2.cell(row=detail_row, column=1, value=s["name"])
                    ws2.cell(row=detail_row, column=2, value="Conversation Notes")
                    ws2.cell(row=detail_row, column=3, value="Direct conversation")
                    ws2.cell(row=detail_row, column=4, value=s.get("conversation_notes", ""))
                    detail_row += 1

                if s.get("conversation_takeaways"):
                    ws2.cell(row=detail_row, column=1, value=s["name"])
                    ws2.cell(row=detail_row, column=2, value="Conversation Takeaways")
                    ws2.cell(row=detail_row, column=3, value="Key takeaways from conversation")
                    ws2.cell(row=detail_row, column=4, value=s.get("conversation_takeaways", ""))
                    detail_row += 1

            ws2.column_dimensions["A"].width = 20
            ws2.column_dimensions["B"].width = 20
            ws2.column_dimensions["C"].width = 50
            ws2.column_dimensions["D"].width = 60

            # --- Sheet 3: Universal Questions ---
            ws3 = wb.create_sheet("Universal Questions")
            uq_headers = ["Speaker", "Date"] + [q["question"] for q in UNIVERSAL_QUESTIONS]
            for col, h in enumerate(uq_headers, 1):
                cell = ws3.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True)

            for row, s in enumerate(speakers, 2):
                ws3.cell(row=row, column=1, value=s["name"])
                ws3.cell(row=row, column=2, value=s["date"])
                ua = s.get("universal_answers", {})
                for col, q in enumerate(UNIVERSAL_QUESTIONS, 3):
                    ws3.cell(row=row, column=col, value=ua.get(q["id"], ""))

            ws3.column_dimensions["A"].width = 20
            ws3.column_dimensions["B"].width = 12
            for col in range(3, 3 + len(UNIVERSAL_QUESTIONS)):
                ws3.column_dimensions[chr(64 + col)].width = 30

            # --- Sheet 4: Company Info ---
            ws4 = wb.create_sheet("Company Info")
            ci_headers = ["Speaker", "Company", "Industry", "Location", "Size", "Founded", "Website", "Notes"]
            for col, h in enumerate(ci_headers, 1):
                cell = ws4.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            for row, s in enumerate(speakers, 2):
                ci = s.get("company_info", {})
                ws4.cell(row=row, column=1, value=s["name"])
                ws4.cell(row=row, column=2, value=s["company"])
                ws4.cell(row=row, column=3, value=ci.get("industry", ""))
                ws4.cell(row=row, column=4, value=ci.get("location", ""))
                ws4.cell(row=row, column=5, value=ci.get("size", ""))
                ws4.cell(row=row, column=6, value=ci.get("founded", ""))
                ws4.cell(row=row, column=7, value=ci.get("website", ""))
                ws4.cell(row=row, column=8, value=ci.get("notes", ""))

            for col in range(1, len(ci_headers) + 1):
                ws4.column_dimensions[chr(64 + col)].width = 20

            # --- Sheet 5: Cross-Speaker Analysis ---
            ws5 = wb.create_sheet("Cross-Speaker Analysis")
            ws5.cell(row=1, column=1, value="Category").font = header_font
            ws5.cell(row=1, column=1).fill = header_fill
            ws5.cell(row=1, column=2, value="Notes").font = header_font
            ws5.cell(row=1, column=2).fill = header_fill

            analysis_items = [
                ("Common Themes on AI Adoption", analysis.get("common_themes", "")),
                ("Biggest Differences in AI Perspective", analysis.get("biggest_differences", "")),
                ("Insights for AI Implementation Consulting", analysis.get("consulting_insights", "")),
                ("Surprises / Changed Assumptions", analysis.get("surprises", ""))
            ]
            for row, (cat, notes) in enumerate(analysis_items, 2):
                ws5.cell(row=row, column=1, value=cat)
                ws5.cell(row=row, column=2, value=notes)

            ws5.column_dimensions["A"].width = 40
            ws5.column_dimensions["B"].width = 80

            # --- Per-Speaker Sheets ---
            for s in speakers:
                sheet_name = s["name"][:31]  # Excel sheet names max 31 chars
                ws = wb.create_sheet(sheet_name)

                # Header info
                ws.cell(row=1, column=1, value="Speaker").font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=2, value=s["name"])
                ws.cell(row=2, column=1, value="Company").font = Font(bold=True)
                ws.cell(row=2, column=2, value=s["company"])
                ws.cell(row=3, column=1, value="Role").font = Font(bold=True)
                ws.cell(row=3, column=2, value=s["role"])
                ws.cell(row=4, column=1, value="Date").font = Font(bold=True)
                ws.cell(row=4, column=2, value=s["date"])
                ws.cell(row=5, column=1, value="Task Reference").font = Font(bold=True)
                ws.cell(row=5, column=2, value=s["task_ref"])

                # Company info
                ci = s.get("company_info", {})
                ws.cell(row=7, column=1, value="Company Information").font = Font(bold=True, size=12)
                ws.cell(row=7, column=1).fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                ws.cell(row=7, column=2).fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                ci_row = 8
                for label, key in [("Industry", "industry"), ("Location", "location"), ("Size", "size"), ("Founded", "founded"), ("Website", "website"), ("Notes", "notes")]:
                    ws.cell(row=ci_row, column=1, value=label).font = Font(bold=True)
                    ws.cell(row=ci_row, column=2, value=ci.get(key, ""))
                    ci_row += 1

                # Shorthand notes
                ci_row += 1
                ws.cell(row=ci_row, column=1, value="Shorthand Notes").font = Font(bold=True, size=12)
                ws.cell(row=ci_row, column=1).fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                ws.cell(row=ci_row, column=2).fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                ci_row += 1
                ws.cell(row=ci_row, column=1, value=s.get("raw_shorthand", ""))
                ci_row += 2

                # Universal questions
                ws.cell(row=ci_row, column=1, value="Universal Questions").font = Font(bold=True, size=12)
                ws.cell(row=ci_row, column=1).fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                ws.cell(row=ci_row, column=2).fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                ci_row += 1
                ua = s.get("universal_answers", {})
                for q in UNIVERSAL_QUESTIONS:
                    ws.cell(row=ci_row, column=1, value=q["question"]).font = Font(bold=True)
                    ws.cell(row=ci_row, column=2, value=ua.get(q["id"], ""))
                    ci_row += 1
                ci_row += 1

                # Sections with prompts and notes
                for section_key, section_name, color in [("ai_usage", "AI Usage", "EBF5FB"), ("takeaways", "Key Takeaways", "FEF9E7"), ("comparisons", "Comparisons", "FDEDEC")]:
                    ws.cell(row=ci_row, column=1, value=section_name).font = Font(bold=True, size=12)
                    ws.cell(row=ci_row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    ws.cell(row=ci_row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    ci_row += 1
                    ws.cell(row=ci_row, column=1, value="Prompt").font = Font(bold=True, italic=True)
                    ws.cell(row=ci_row, column=2, value="Notes").font = Font(bold=True, italic=True)
                    ci_row += 1
                    for item in s.get(section_key, []):
                        ws.cell(row=ci_row, column=1, value=item.get("prompt", ""))
                        ws.cell(row=ci_row, column=2, value=item.get("notes", ""))
                        ci_row += 1
                    ci_row += 1

                # Conversation notes if applicable
                if s.get("has_conversation_task"):
                    ws.cell(row=ci_row, column=1, value="Conversation Notes").font = Font(bold=True, size=12)
                    ws.cell(row=ci_row, column=1).fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
                    ws.cell(row=ci_row, column=2).fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
                    ci_row += 1
                    ws.cell(row=ci_row, column=1, value="Notes").font = Font(bold=True)
                    ws.cell(row=ci_row, column=2, value=s.get("conversation_notes", ""))
                    ci_row += 1
                    ws.cell(row=ci_row, column=1, value="Key Takeaways").font = Font(bold=True)
                    ws.cell(row=ci_row, column=2, value=s.get("conversation_takeaways", ""))
                    ci_row += 1

                ws.column_dimensions["A"].width = 45
                ws.column_dimensions["B"].width = 65

            # Save to buffer and download
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="📥 Download Excel File",
                data=buffer,
                file_name=f"SFE_Speaker_Notes_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Excel file ready for download!")

        except ImportError:
            st.error("openpyxl not installed. Add 'openpyxl' to requirements.txt and redeploy.")

# --- Add Speaker ---
elif selection == "➕ Add Speaker":
    st.title("Add New Speaker")

    with st.form("add_speaker"):
        name = st.text_input("Speaker Name")
        date = st.date_input("Date of Talk")
        company = st.text_input("Company / Organization")
        role = st.text_input("Role / Title")
        bio = st.text_area("Background / Bio (optional)")
        has_conv = st.checkbox("This speaker has a conversation/interview task")

        st.markdown("**Company Information**")
        c_industry = st.text_input("Industry")
        c_location = st.text_input("Location")
        c_size = st.text_input("Company Size")
        c_website = st.text_input("Website")
        c_notes = st.text_area("Company Notes")

        submitted = st.form_submit_button("Add Speaker")

        if submitted and name:
            new_id = name.lower().replace(" ", "_")
            new_speaker = {
                "id": new_id,
                "name": name,
                "date": date.strftime("%Y-%m-%d"),
                "company": company,
                "role": role,
                "task_ref": "Task 1b (running doc)" + (" + conversation task" if has_conv else ""),
                "bio": bio,
                "company_info": {
                    "industry": c_industry,
                    "location": c_location,
                    "size": c_size,
                    "founded": "",
                    "website": c_website,
                    "notes": c_notes
                },
                "has_conversation_task": has_conv,
                "ai_usage": [
                    {"prompt": "How does this speaker/organization use AI?", "notes": ""},
                    {"prompt": "What tools or platforms do they rely on?", "notes": ""},
                    {"prompt": "What is their attitude toward AI adoption?", "notes": ""},
                    {"prompt": "Biggest AI-related challenge they face", "notes": ""}
                ],
                "takeaways": [
                    {"prompt": "Most valuable insight from the talk", "notes": ""},
                    {"prompt": "Advice relevant to my career path", "notes": ""},
                    {"prompt": "Anything surprising or unexpected", "notes": ""}
                ],
                "comparisons": [
                    {"prompt": "How does this speaker's perspective compare to previous speakers?", "notes": ""},
                    {"prompt": "What new pattern or contradiction emerged?", "notes": ""}
                ],
                "prepared_questions": [],
                "conversation_notes": "",
                "conversation_takeaways": "",
                "raw_shorthand": "",
                "universal_answers": {}
            }
            speakers.append(new_speaker)
            save_data({"speakers": speakers, "analysis": analysis})
            st.success(f"Added {name}!")
            st.rerun()

# --- Cross-Speaker Analysis ---
elif selection == "🔍 Cross-Speaker Analysis":
    st.title("Cross-Speaker Analysis")
    st.markdown("Compare and analyze patterns across all speakers. Build this throughout the quarter.")
    st.markdown("---")

    sections = [
        ("common_themes", "Common Themes on AI Adoption",
         "What patterns emerged? Where did multiple speakers agree? Common barriers mentioned (cost, trust, skills, culture)?"),
        ("biggest_differences", "Biggest Differences in AI Perspective",
         "Where did speakers disagree? How much did industry/company size drive differences? Enterprise vs. small business vs. advisory?"),
        ("consulting_insights", "Insights Relevant to AI Implementation Consulting",
         "What did speakers teach you about what companies actually need? What skills kept coming up? How does this connect to failure rate data from Task 1a research?"),
        ("surprises", "Surprises / Changed Assumptions",
         "What challenged something you believed? Biggest gap between expectation and reality? How has your understanding evolved?")
    ]

    values = {}
    for key, title, description in sections:
        st.subheader(title)
        st.markdown(f"*{description}*")
        values[key] = st.text_area(key, value=analysis.get(key, ""), height=150, key=f"analysis_{key}", label_visibility="collapsed")
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    if st.button("Save Analysis", type="primary"):
        for key in values:
            analysis[key] = values[key]
        save_data({"speakers": speakers, "analysis": analysis})
        st.success("Analysis saved.")

# --- Speaker Pages ---
else:
    speaker_name = selection.split(" ", 1)[1]
    speaker = None
    speaker_idx = None
    for i, s in enumerate(speakers):
        if s["name"] == speaker_name:
            speaker = s
            speaker_idx = i
            break

    if speaker is None:
        st.error("Speaker not found.")
    else:
        speaker_date = datetime.strptime(speaker["date"], "%Y-%m-%d")
        is_upcoming = speaker_date > datetime.now()
        filled, total_prompts = count_notes(speaker)

        # Header
        st.markdown(f"""
        <div class="speaker-header">
            <h2>{speaker['name']}</h2>
            <p><strong>{speaker['company']}</strong> — {speaker['role']}</p>
            <p>{speaker_date.strftime('%B %d, %Y')} &nbsp;|&nbsp; {speaker['task_ref']}</p>
            <p>Notes: {filled}/{total_prompts} filled</p>
        </div>
        """, unsafe_allow_html=True)

        # Bio
        if speaker.get("bio"):
            st.markdown(f"*{speaker['bio']}*")

        # --- Company Information Card ---
        ci = speaker.get("company_info", {})
        st.markdown(f"""
        <div class="company-card">
            <h4>🏢 Company Information</h4>
            <div class="info-row"><span class="info-label">Industry:</span> {ci.get('industry', '')}</div>
            <div class="info-row"><span class="info-label">Location:</span> {ci.get('location', '')}</div>
            <div class="info-row"><span class="info-label">Size:</span> {ci.get('size', '')}</div>
            <div class="info-row"><span class="info-label">Founded:</span> {ci.get('founded', '')}</div>
            <div class="info-row"><span class="info-label">Website:</span> {ci.get('website', '')}</div>
            <div class="info-row"><span class="info-label">Notes:</span> {ci.get('notes', '')}</div>
        </div>
        """, unsafe_allow_html=True)

        # Edit company info
        with st.expander("✏️ Edit Company Information"):
            ci_industry = st.text_input("Industry", value=ci.get("industry", ""), key=f"ci_ind_{speaker['id']}")
            ci_location = st.text_input("Location", value=ci.get("location", ""), key=f"ci_loc_{speaker['id']}")
            ci_size = st.text_input("Size", value=ci.get("size", ""), key=f"ci_size_{speaker['id']}")
            ci_founded = st.text_input("Founded", value=ci.get("founded", ""), key=f"ci_found_{speaker['id']}")
            ci_website = st.text_input("Website", value=ci.get("website", ""), key=f"ci_web_{speaker['id']}")
            ci_notes = st.text_area("Company Notes", value=ci.get("notes", ""), key=f"ci_notes_{speaker['id']}")
            if st.button("Update Company Info", key=f"ci_save_{speaker['id']}"):
                speakers[speaker_idx]["company_info"] = {
                    "industry": ci_industry, "location": ci_location, "size": ci_size,
                    "founded": ci_founded, "website": ci_website, "notes": ci_notes
                }
                save_data({"speakers": speakers, "analysis": analysis})
                st.success("Company info updated!")
                st.rerun()

        st.markdown("---")

        # --- Shorthand Notes ---
        st.subheader("📝 Live Shorthand Notes")
        st.markdown("*Quick notes during the talk. Keywords, phrases, ideas. Fill in structured sections after.*")
        raw = st.text_area("shorthand", value=speaker.get("raw_shorthand", ""), height=120,
                           key=f"raw_{speaker['id']}", placeholder="Type quick notes during the talk...",
                           label_visibility="collapsed")

        st.markdown("---")

        # --- Universal Questions (cross-speaker comparison dataset) ---
        st.subheader("📊 Universal Questions")
        st.markdown("*Same questions for every speaker. Builds a dataset for cross-speaker comparison and Task 8 analysis.*")
        
        universal = speaker.get("universal_answers", {})
        uq_values = {}
        for q in UNIVERSAL_QUESTIONS:
            badge = "🔢" if q["type"] == "quantitative" else "🏷️"
            st.markdown(f'<div class="prompt-box">{badge} {q["question"]}</div>', unsafe_allow_html=True)
            val = st.text_input(
                q["id"], value=universal.get(q["id"], ""),
                key=f"uq_{speaker['id']}_{q['id']}",
                label_visibility="collapsed",
                placeholder="Their answer / your observation..."
            )
            uq_values[q["id"]] = val

        st.markdown("---")

        # --- Prepared Questions ---
        if speaker.get("has_conversation_task") or speaker.get("prepared_questions"):
            st.subheader("❓ Prepared Questions")
            for i, q in enumerate(speaker.get("prepared_questions", [])):
                st.markdown(f'<div class="prompt-box">• {q}</div>', unsafe_allow_html=True)
            new_q = st.text_input("Add a question", key=f"newq_{speaker['id']}",
                                  placeholder="Type a new question and hit Enter...")
            if new_q:
                speakers[speaker_idx]["prepared_questions"].append(new_q)
                save_data({"speakers": speakers, "analysis": analysis})
                st.rerun()
            st.markdown("---")

        # --- AI Usage ---
        st.subheader("🤖 How They Apply / See AI in Their Business")
        ai_values = []
        for i, item in enumerate(speaker.get("ai_usage", [])):
            st.markdown(f'<div class="prompt-box">{item["prompt"]}</div>', unsafe_allow_html=True)
            val = st.text_area(f"ai_note_{i}", value=item.get("notes", ""), height=80,
                               key=f"ai_{speaker['id']}_{i}", label_visibility="collapsed",
                               placeholder="Your notes on this...")
            ai_values.append(val)

        st.markdown("---")

        # --- Key Takeaways ---
        st.subheader("💡 Key Takeaways / General Notes")
        takeaway_values = []
        for i, item in enumerate(speaker.get("takeaways", [])):
            st.markdown(f'<div class="prompt-box">{item["prompt"]}</div>', unsafe_allow_html=True)
            val = st.text_area(f"take_note_{i}", value=item.get("notes", ""), height=80,
                               key=f"take_{speaker['id']}_{i}", label_visibility="collapsed",
                               placeholder="Your notes on this...")
            takeaway_values.append(val)

        st.markdown("---")

        # --- Conversation Notes ---
        if speaker.get("has_conversation_task"):
            st.subheader("🗣️ Conversation Notes")
            conv_notes = st.text_area("conv", value=speaker.get("conversation_notes", ""), height=150,
                                      key=f"conv_{speaker['id']}", label_visibility="collapsed",
                                      placeholder="What did you discuss? Key points from the conversation...")

            st.subheader("🎯 Key Takeaways from Conversation")
            conv_takeaways = st.text_area("convtake", value=speaker.get("conversation_takeaways", ""), height=120,
                                          key=f"convtake_{speaker['id']}", label_visibility="collapsed",
                                          placeholder="Most important thing learned, how it changes your thinking...")
            st.markdown("---")

        # --- Comparisons ---
        st.subheader("🔄 Comparison / Contrast with Other Speakers")
        comp_values = []
        for i, item in enumerate(speaker.get("comparisons", [])):
            st.markdown(f'<div class="prompt-box">{item["prompt"]}</div>', unsafe_allow_html=True)
            val = st.text_area(f"comp_note_{i}", value=item.get("notes", ""), height=80,
                               key=f"comp_{speaker['id']}_{i}", label_visibility="collapsed",
                               placeholder="Your notes on this...")
            comp_values.append(val)

        # --- Save All ---
        st.markdown("---")
        if st.button("💾 Save All Notes", type="primary", key=f"save_{speaker['id']}"):
            speakers[speaker_idx]["raw_shorthand"] = raw
            speakers[speaker_idx]["universal_answers"] = uq_values
            for i, val in enumerate(ai_values):
                speakers[speaker_idx]["ai_usage"][i]["notes"] = val
            for i, val in enumerate(takeaway_values):
                speakers[speaker_idx]["takeaways"][i]["notes"] = val
            for i, val in enumerate(comp_values):
                speakers[speaker_idx]["comparisons"][i]["notes"] = val
            if speaker.get("has_conversation_task"):
                speakers[speaker_idx]["conversation_notes"] = conv_notes
                speakers[speaker_idx]["conversation_takeaways"] = conv_takeaways
            save_data({"speakers": speakers, "analysis": analysis})
            st.success("All notes saved!")
